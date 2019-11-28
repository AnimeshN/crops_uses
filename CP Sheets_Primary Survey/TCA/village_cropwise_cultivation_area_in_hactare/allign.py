# -*- coding: utf-8 -*-
import pandas as pd


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


directory = pd.read_csv('dictionary.csv')
years = [2018,2013,2008,2003,1998]

output = pd.read_excel('Master.xlsx', sheet_name=None)
dfObj = output['Sheet1']

listofseries = []
for ind in directory.index: 
    
    village_name = directory['Village'][ind]
    file = directory['filename'][ind]

    df = pd.read_excel(file+'.xlsx', sheet_name=None)

    my_df = df['2018'][['खरीप प्रमुख पिके','total_cultivable_area']].values.tolist()
    length = len(my_df)

    for i in range(length):
        if my_df[i][0] == 'खरीप भाजीपाला पिके':
            a = i
        if my_df[i][0] == 'दीर्घ खरीप पिके':
            b = i
        if my_df[i][0] == 'रब्बी प्रमुख पिके':
            c = i
        if my_df[i][0] == 'रब्बी भाजीपाला पिके':
            d = i
        if my_df[i][0] == 'उन्हाळी प्रमुख पिके':
            e = i
        if my_df[i][0] == 'उन्हाळी भाजीपाला पिके':
            f = i
        if my_df[i][0] == 'वार्षिक व बहुवर्षीय पिके':
            g = i

    K1,K2,K3,R4,R5,S6,S7,A8 = [],[],[],[],[],[],[],[]
    for i in range(length):
        if i < a:
            K1.append([my_df[i][0],my_df[i][1]])
        elif (i > a and i < b):
            K2.append([my_df[i][0],my_df[i][1]])
        elif (i > b and i < c):
            K3.append([my_df[i][0],my_df[i][1]])
        elif (i > c and i < d):
            R4.append([my_df[i][0],my_df[i][1]])
        elif (i > d and i < e):
            R5.append([my_df[i][0],my_df[i][1]])
        elif (i > e and i < f):
            S6.append([my_df[i][0],my_df[i][1]])
        elif (i > f and i < g):
            S7.append([my_df[i][0],my_df[i][1]])
        elif (i > g):
            A8.append([my_df[i][0],my_df[i][1]])

    KHARIF = [['खरीप ज्वारी'],['बाजरी'],['मुग'],['उडिद'],['तीळ'],['सोयाबीन'],['मका'],['खरीप भुईमुग'],['खरीप सुर्यफुल'],['खरीप चारा पिके'],['खरीप भात'],['नागली'],['वरई'],['वाल'],['Others']]
    mymaster = KHARIF

    for i in range(len(K1)):
    #     print(K1[i][0])
        for j in range(len(mymaster)):
    #         print(mymaster[j][0])
            if K1[i][0] == mymaster[j][0]:
                mymaster[j].append(K1[i][1])
                break
    #     print(mymaster)
    # print(mymaster)

    listlist = [village_name]
    total = 0
    for i in range(len(mymaster)):
        if len(mymaster[i]) == 2:
            listlist.append(mymaster[i][1])
            total = total + mymaster[i][1]
        else:
            listlist.append("")
    listlist.append(total) 

    series = pd.Series(listlist)
    listofseries.append(series)


modDfObj = dfObj.append(listofseries)
print(listofseries)
# modDfObj.to_excel("OUT/master_kharif.xlsx")


        

