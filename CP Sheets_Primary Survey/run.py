# -*- coding: utf-8 -*-
import pandas as pd
import sys
inputfile = sys.argv[1]
TCA = float(sys.argv[2])
print(inputfile)
dfs = pd.read_excel("./"+inputfile+'.xlsx',header=[1], sheet_name=None)
cp2018,cp2013,cp2008,cp2003,cp1998 = dfs['CPData-2018'],dfs['CPData-2013'],dfs['CPData-2008'],dfs['CPData-2003'],dfs['CPData-1998']

cp2018 = cp2018.dropna(how="all")
cp2013 = cp2013.dropna(how="all")
cp2008 = cp2008.dropna(how="all")
cp2003 = cp2003.dropna(how="all")
cp1998 = cp1998.dropna(how="all")

total_land = cp2018['Total Land Holding'].sum()

col = ['Interview Number','Farmer Name','Contact No.','Gat No.','Total Land Holding','Rainfed','Irrigated','पडिक जमीन']
cp2018.loc[:,col] = cp2018.loc[:,col].ffill()
cp2013.loc[:,col] = cp2013.loc[:,col].ffill()
cp2008.loc[:,col] = cp2008.loc[:,col].ffill()
cp2003.loc[:,col] = cp2003.loc[:,col].ffill()
cp1998.loc[:,col] = cp1998.loc[:,col].ffill()

req_col = ['Interview Number', 'Farmer Name','Total Land Holding','पडिक जमीन','खरीप प्रमुख पिके', 'Area (acres)', 'खरीप भाजीपाला पिके','Area (acres).1', 'दीर्घ खरीप पिके', 'Area (acres).2','रब्बी प्रमुख पिके', 'Area (acres).3', 'रब्बी भाजीपाला पिके','Area (acres).4', 'उन्हाळी प्रमुख पिके', 'Area (acres).5','उन्हाळी भाजीपाला पिके', 'Area (acres).6', 'वार्षिक व बहुवर्षीय पिके','Area (acres).7']
df2018 = cp2018[req_col]
df2013 = cp2013[req_col]
df2008 = cp2008[req_col]
df2003 = cp2003[req_col]
df1998 = cp1998[req_col]


col_names = ['Interview Number','Farmer Name','Total Land Holding','पडिक जमीन','खरीप प्रमुख पिके','area_kharif_major_1','खरीप भाजीपाला पिके','area_kharif_vegi_2','दीर्घ खरीप पिके','area_Kharif_long_3','रब्बी प्रमुख पिके','area_rabi_main_4','रब्बी भाजीपाला पिके','area_rabi_vegi_5','उन्हाळी प्रमुख पिके','area_summer_major_6','उन्हाळी भाजीपाला पिके','area_summer_vegi_7','वार्षिक व बहुवर्षीय पिके','area_annual_8']
df2018.columns = col_names
df2013.columns = col_names
df2008.columns = col_names
df2003.columns = col_names
df1998.columns = col_names

k1_18=df2018.groupby(['खरीप प्रमुख पिके']).agg({
    'area_kharif_major_1':sum,
})
if not k1_18.empty:
    k1_18['crop_area_per_holding_sample'] = k1_18.apply(lambda row: round(row.area_kharif_major_1/total_land,4), axis = 1)
    k1_18['total_cultivable_area'] = k1_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    

k1_13=df2013.groupby(['खरीप प्रमुख पिके']).agg({
    'area_kharif_major_1':sum,
})
if not k1_13.empty:
    k1_13['crop_area_per_holding_sample'] = k1_13.apply(lambda row: round(row.area_kharif_major_1/total_land,4), axis = 1)
    k1_13['total_cultivable_area'] = k1_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)


k1_08=df2008.groupby(['खरीप प्रमुख पिके']).agg({
    'area_kharif_major_1':sum,
})
# k1_08['total_land'] = k1_08.apply(lambda row: total_land, axis = 1)
if not k1_08.empty:
    k1_08['crop_area_per_holding_sample'] = k1_08.apply(lambda row: round(row.area_kharif_major_1/total_land,4), axis = 1)
    k1_08['total_cultivable_area'] = k1_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k1_03=df2003.groupby(['खरीप प्रमुख पिके']).agg({
    'area_kharif_major_1':sum,
})
# k1_03['total_land'] = k1_03.apply(lambda row: total_land, axis = 1)
if not k1_03.empty:
    k1_03['crop_area_per_holding_sample'] = k1_03.apply(lambda row: round(row.area_kharif_major_1/total_land,4), axis = 1) 
    k1_03['total_cultivable_area'] = k1_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

k1_98=df1998.groupby(['खरीप प्रमुख पिके']).agg({
    'area_kharif_major_1':sum,
})
# k1_98['total_land'] = k1_98.apply(lambda row: total_land, axis = 1)
if not k1_98.empty:
    k1_98['crop_area_per_holding_sample'] = k1_98.apply(lambda row: round(row.area_kharif_major_1/total_land,4), axis = 1) 
    k1_98['total_cultivable_area'] = k1_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

k2_18=df2018.groupby(['खरीप भाजीपाला पिके']).agg({
    'area_kharif_vegi_2':sum,
})
# k2_18['total_land'] = k2_18.apply(lambda row: total_land, axis = 1)
if not k2_18.empty:
    k2_18['crop_area_per_holding_sample'] = k2_18.apply(lambda row: round(row.area_kharif_vegi_2/total_land,4), axis = 1)
    k2_18['total_cultivable_area'] = k2_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k2_13=df2013.groupby(['खरीप भाजीपाला पिके']).agg({
    'area_kharif_vegi_2':sum,
})
# k2_13['total_land'] = k2_13.apply(lambda row: total_land, axis = 1)
if not k2_13.empty:
    k2_13['crop_area_per_holding_sample'] = k2_13.apply(lambda row: round(row.area_kharif_vegi_2/total_land,4), axis = 1)
    k2_13['total_cultivable_area'] = k2_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k2_08=df2008.groupby(['खरीप भाजीपाला पिके']).agg({
    'area_kharif_vegi_2':sum,
})
# k2_08['total_land'] = k2_08.apply(lambda row: total_land, axis = 1)
if not k2_08.empty:
    k2_08['crop_area_per_holding_sample'] = k2_08.apply(lambda row: round(row.area_kharif_vegi_2/total_land,4), axis = 1)
    k2_08['total_cultivable_area'] = k2_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k2_03=df2003.groupby(['खरीप भाजीपाला पिके']).agg({
    'area_kharif_vegi_2':sum,
})
# k2_03['total_land'] = k2_03.apply(lambda row: total_land, axis = 1)
if not k2_03.empty:
    k2_03['crop_area_per_holding_sample'] = k2_03.apply(lambda row: round(row.area_kharif_vegi_2/total_land,4), axis = 1) 
    k2_03['total_cultivable_area'] = k2_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

k2_98=df1998.groupby(['खरीप भाजीपाला पिके']).agg({
    'area_kharif_vegi_2':sum,
})
# k2_98['total_land'] = k2_98.apply(lambda row: total_land, axis = 1)
if not k2_98.empty:
    k2_98['crop_area_per_holding_sample'] = k2_98.apply(lambda row: round(row.area_kharif_vegi_2/total_land,4), axis = 1) 
    k2_98['total_cultivable_area'] = k2_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k3_18=df2018.groupby(['दीर्घ खरीप पिके']).agg({
    'area_Kharif_long_3':sum,
})
# k3_18['total_land'] = k3_18.apply(lambda row: total_land, axis = 1)
if not k3_18.empty:
    k3_18['crop_area_per_holding_sample'] = k3_18.apply(lambda row: round(row.area_Kharif_long_3/total_land,4), axis = 1)
    k3_18['total_cultivable_area'] = k3_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k3_13=df2013.groupby(['दीर्घ खरीप पिके']).agg({
    'area_Kharif_long_3':sum,
})
# k3_13['total_land'] = k3_13.apply(lambda row: total_land, axis = 1)
if not k3_13.empty:
    k3_13['crop_area_per_holding_sample'] = k3_13.apply(lambda row: round(row.area_Kharif_long_3/total_land,4), axis = 1)
    k3_13['total_cultivable_area'] = k3_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k3_08=df2008.groupby(['दीर्घ खरीप पिके']).agg({
    'area_Kharif_long_3':sum,
})
# k3_08['total_land'] = k3_08.apply(lambda row: total_land, axis = 1)
if not k3_08.empty:
    k3_08['crop_area_per_holding_sample'] = k3_08.apply(lambda row: round(row.area_Kharif_long_3/total_land,4), axis = 1)
    k3_08['total_cultivable_area'] = k3_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
k3_03=df2003.groupby(['दीर्घ खरीप पिके']).agg({
    'area_Kharif_long_3':sum,
})
# k3_03['total_land'] = k3_03.apply(lambda row: total_land, axis = 1)
if not k3_03.empty:
    k3_03['crop_area_per_holding_sample'] = k3_03.apply(lambda row: round(row.area_Kharif_long_3/total_land,4), axis = 1) 
    k3_03['total_cultivable_area'] = k3_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

k3_98=df1998.groupby(['दीर्घ खरीप पिके']).agg({
    'area_Kharif_long_3':sum,
})
# k3_98['total_land'] = k3_98.apply(lambda row: total_land, axis = 1)
if not k3_98.empty:
    k3_98['crop_area_per_holding_sample'] = k3_98.apply(lambda row: round(row.area_Kharif_long_3/total_land,4), axis = 1) 
    k3_98['total_cultivable_area'] = k3_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
r4_18=df2018.groupby(['रब्बी प्रमुख पिके']).agg({
    'area_rabi_main_4':sum,
})
# r4_18['total_land'] = r4_18.apply(lambda row: total_land, axis = 1)
if not r4_18.empty:
    r4_18['crop_area_per_holding_sample'] = r4_18.apply(lambda row: round(row.area_rabi_main_4/total_land,4), axis = 1)
    r4_18['total_cultivable_area'] = r4_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
r4_13=df2013.groupby(['रब्बी प्रमुख पिके']).agg({
    'area_rabi_main_4':sum,
})
# r4_13['total_land'] = r4_13.apply(lambda row: total_land, axis = 1)
if not r4_13.empty:
    r4_13['crop_area_per_holding_sample'] = r4_13.apply(lambda row: round(row.area_rabi_main_4/total_land,4), axis = 1)
    r4_13['total_cultivable_area'] = r4_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
r4_08=df2008.groupby(['रब्बी प्रमुख पिके']).agg({
    'area_rabi_main_4':sum,
})
# r4_08['total_land'] = r4_08.apply(lambda row: total_land, axis = 1)
if not r4_08.empty:
    r4_08['crop_area_per_holding_sample'] = r4_08.apply(lambda row: round(row.area_rabi_main_4/total_land,4), axis = 1)
    r4_08['total_cultivable_area'] = r4_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
    
r4_03=df2003.groupby(['रब्बी प्रमुख पिके']).agg({
    'area_rabi_main_4':sum,
})
# r4_03['total_land'] = r4_03.apply(lambda row: total_land, axis = 1)
if not r4_03.empty:
    r4_03['crop_area_per_holding_sample'] = r4_03.apply(lambda row: round(row.area_rabi_main_4/total_land,4), axis = 1) 
    r4_03['total_cultivable_area'] = r4_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

r4_98=df1998.groupby(['रब्बी प्रमुख पिके']).agg({
    'area_rabi_main_4':sum,
})
# r4_98['total_land'] = r4_98.apply(lambda row: total_land, axis = 1)
if not r4_98.empty:
    r4_98['crop_area_per_holding_sample'] = r4_98.apply(lambda row: round(row.area_rabi_main_4/total_land,4), axis = 1) 
    r4_98['total_cultivable_area'] = r4_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)


r5_18=df2018.groupby(['रब्बी भाजीपाला पिके']).agg({
    'area_rabi_vegi_5':sum,
})
# r5_18['total_land'] = r5_18.apply(lambda row: total_land, axis = 1)
if not r5_18.empty:
    r5_18['crop_area_per_holding_sample'] = r5_18.apply(lambda row: round(row.area_rabi_vegi_5/total_land,4), axis = 1)
    r5_18['total_cultivable_area'] = r5_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
r5_13=df2013.groupby(['रब्बी भाजीपाला पिके']).agg({
    'area_rabi_vegi_5':sum,
})
# r5_13['total_land'] = r5_13.apply(lambda row: total_land, axis = 1)
if not r5_13.empty:
    r5_13['crop_area_per_holding_sample'] = r5_13.apply(lambda row: round(row.area_rabi_vegi_5/total_land,4), axis = 1)
    r5_13['total_cultivable_area'] = r5_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
r5_08=df2008.groupby(['रब्बी भाजीपाला पिके']).agg({
    'area_rabi_vegi_5':sum,
})
# r5_08['total_land'] = r5_08.apply(lambda row: total_land, axis = 1)
if not r5_08.empty:
    r5_08['crop_area_per_holding_sample'] = r5_08.apply(lambda row: round(row.area_rabi_vegi_5/total_land,4), axis = 1)
    r5_08['total_cultivable_area'] = r5_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
r5_03=df2003.groupby(['रब्बी भाजीपाला पिके']).agg({
    'area_rabi_vegi_5':sum,
})
# r5_03['total_land'] = r5_03.apply(lambda row: total_land, axis = 1)
if not r5_03.empty:
    r5_03['crop_area_per_holding_sample'] = r5_03.apply(lambda row: round(row.area_rabi_vegi_5/total_land,4), axis = 1) 
    r5_03['total_cultivable_area'] = r5_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
r5_98=df1998.groupby(['रब्बी भाजीपाला पिके']).agg({
    'area_rabi_vegi_5':sum,
})
# r5_98['total_land'] = r5_98.apply(lambda row: total_land, axis = 1)
if not r5_98.empty:
    r5_98['crop_area_per_holding_sample'] = r5_98.apply(lambda row: round(row.area_rabi_vegi_5/total_land,4), axis = 1) 
    r5_98['total_cultivable_area'] = r5_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    

s6_18=df2018.groupby(['उन्हाळी प्रमुख पिके']).agg({
    'area_summer_major_6':sum,
})
# s6_18['total_land'] = s6_18.apply(lambda row: total_land, axis = 1)
if not s6_18.empty:
    s6_18['crop_area_per_holding_sample'] = s6_18.apply(lambda row: round(row.area_summer_major_6/total_land,4), axis = 1)
    s6_18['total_cultivable_area'] = s6_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)        
s6_13=df2013.groupby(['उन्हाळी प्रमुख पिके']).agg({
    'area_summer_major_6':sum,
})
# s6_13['total_land'] = s6_13.apply(lambda row: total_land, axis = 1)
if not s6_13.empty:
    s6_13['crop_area_per_holding_sample'] = s6_13.apply(lambda row: round(row.area_summer_major_6/total_land,4), axis = 1)
    s6_13['total_cultivable_area'] = s6_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)        
s6_08=df2008.groupby(['उन्हाळी प्रमुख पिके']).agg({
    'area_summer_major_6':sum,
})
# s6_08['total_land'] = s6_08.apply(lambda row: total_land, axis = 1)
if not s6_08.empty:
    s6_08['crop_area_per_holding_sample'] = s6_08.apply(lambda row: round(row.area_summer_major_6/total_land,4), axis = 1)
    s6_08['total_cultivable_area'] = s6_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)        
s6_03=df2003.groupby(['उन्हाळी प्रमुख पिके']).agg({
    'area_summer_major_6':sum,
})
# s6_03['total_land'] = s6_03.apply(lambda row: total_land, axis = 1)
if not s6_03.empty:
    s6_03['crop_area_per_holding_sample'] = s6_03.apply(lambda row: round(row.area_summer_major_6/total_land,4), axis = 1) 
    s6_03['total_cultivable_area'] = s6_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
s6_98=df1998.groupby(['उन्हाळी प्रमुख पिके']).agg({
    'area_summer_major_6':sum,
})
# s6_98['total_land'] = s6_98.apply(lambda row: total_land, axis = 1)
if not s6_98.empty:
    s6_98['crop_area_per_holding_sample'] = s6_98.apply(lambda row: round(row.area_summer_major_6/total_land,4), axis = 1) 
    s6_98['total_cultivable_area'] = s6_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

s7_18=df2018.groupby(['उन्हाळी भाजीपाला पिके']).agg({
    'area_summer_vegi_7':sum,
})
# s7_18['total_land'] = s7_18.apply(lambda row: total_land, axis = 1)
if not s7_18.empty:
    s7_18['crop_area_per_holding_sample'] = s7_18.apply(lambda row: round(row.area_summer_vegi_7/total_land,4), axis = 1)
    s7_18['total_cultivable_area'] = s7_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
s7_13=df2013.groupby(['उन्हाळी भाजीपाला पिके']).agg({
    'area_summer_vegi_7':sum,
})
# s7_13['total_land'] = s7_13.apply(lambda row: total_land, axis = 1)
if not s7_13.empty:
    s7_13['crop_area_per_holding_sample'] = s7_13.apply(lambda row: round(row.area_summer_vegi_7/total_land,4), axis = 1)
    s7_13['total_cultivable_area'] = s7_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
s7_08=df2008.groupby(['उन्हाळी भाजीपाला पिके']).agg({
    'area_summer_vegi_7':sum,
})
# s7_08['total_land'] = s7_08.apply(lambda row: total_land, axis = 1)
if not s7_08.empty:
    s7_08['crop_area_per_holding_sample'] = s7_08.apply(lambda row: round(row.area_summer_vegi_7/total_land,4), axis = 1)
    s7_08['total_cultivable_area'] = s7_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
s7_03=df2003.groupby(['उन्हाळी भाजीपाला पिके']).agg({
    'area_summer_vegi_7':sum,
})
# s7_03['total_land'] = s7_03.apply(lambda row: total_land, axis = 1)
if not s7_03.empty:
    s7_03['crop_area_per_holding_sample'] = s7_03.apply(lambda row: round(row.area_summer_vegi_7/total_land,4), axis = 1) 
    s7_03['total_cultivable_area'] = s7_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
s7_98=df1998.groupby(['उन्हाळी भाजीपाला पिके']).agg({
    'area_summer_vegi_7':sum,
})
# s7_98['total_land'] = s7_98.apply(lambda row: total_land, axis = 1)
if not s7_98.empty:
    s7_98['crop_area_per_holding_sample'] = s7_98.apply(lambda row: round(row.area_summer_vegi_7/total_land,4), axis = 1) 
    s7_98['total_cultivable_area'] = s7_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)

a8_18=df2018.groupby(['वार्षिक व बहुवर्षीय पिके']).agg({
    'area_annual_8':sum,
})
# a8_18['total_land'] = a8_18.apply(lambda row: total_land, axis = 1)
if not a8_18.empty:
    a8_18['crop_area_per_holding_sample'] = a8_18.apply(lambda row: round(row.area_annual_8/total_land,4), axis = 1)
    a8_18['total_cultivable_area'] = a8_18.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
a8_13=df2013.groupby(['वार्षिक व बहुवर्षीय पिके']).agg({
    'area_annual_8':sum,
})
# a8_13['total_land'] = a8_13.apply(lambda row: total_land, axis = 1)
if not a8_13.empty:
    a8_13['crop_area_per_holding_sample'] = a8_13.apply(lambda row: round(row.area_annual_8/total_land,4), axis = 1)
    a8_13['total_cultivable_area'] = a8_13.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
a8_08=df2008.groupby(['वार्षिक व बहुवर्षीय पिके']).agg({
    'area_annual_8':sum,
})
# a8_08['total_land'] = a8_08.apply(lambda row: total_land, axis = 1)
if not a8_08.empty:
    a8_08['crop_area_per_holding_sample'] = a8_08.apply(lambda row: round(row.area_annual_8/total_land,4), axis = 1)
    a8_08['total_cultivable_area'] = a8_08.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)    
a8_03=df2003.groupby(['वार्षिक व बहुवर्षीय पिके']).agg({
    'area_annual_8':sum,
})
# a8_03['total_land'] = a8_03.apply(lambda row: total_land, axis = 1)
if not a8_03.empty:
    a8_03['crop_area_per_holding_sample'] = a8_03.apply(lambda row: round(row.area_annual_8/total_land,4), axis = 1) 
    a8_03['total_cultivable_area'] = a8_03.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)
a8_98=df1998.groupby(['वार्षिक व बहुवर्षीय पिके']).agg({
    'area_annual_8':sum,
})
# a8_98['total_land'] = a8_98.apply(lambda row: total_land, axis = 1)
if not a8_98.empty:
    a8_98['crop_area_per_holding_sample'] = a8_98.apply(lambda row: round(row.area_annual_8/total_land,4), axis = 1) 
    a8_98['total_cultivable_area'] = a8_98.apply(lambda row: row.crop_area_per_holding_sample*TCA, axis = 1)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
   
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


#18
k1_18.to_excel("TCA/"+inputfile+"_TCA.xlsx",sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k2_18,sheet_name='2018')
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k3_18,sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r4_18,sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r5_18,sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s6_18,sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s7_18,sheet_name="2018")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",a8_18,sheet_name="2018")

#13
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k1_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k2_13,sheet_name='2013')
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k3_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r4_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r5_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s6_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s7_13,sheet_name="2013")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",a8_13,sheet_name="2013")

#08
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k1_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k2_08,sheet_name='2008')
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k3_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r4_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r5_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s6_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s7_08,sheet_name="2008")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",a8_08,sheet_name="2008")

#03
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k1_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k2_03,sheet_name='2003')
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k3_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r4_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r5_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s6_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s7_03,sheet_name="2003")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",a8_03,sheet_name="2003")

#98
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k1_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k2_98,sheet_name='1998')
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",k3_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r4_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",r5_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s6_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",s7_98,sheet_name="1998")
append_df_to_excel("TCA/"+inputfile+"_TCA.xlsx",a8_98,sheet_name="1998")
